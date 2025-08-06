import time
import os
import re
import json
import ipaddress
import csv
from datetime import datetime 
import pandas as pd
from openpyxl import load_workbook

"""
    该py文件存放了AutoDev_SheetTools类和AutoDev_OtherTools类，实例化对象时一般缩写为ADST和ADOT。
    AutoDev_SheetTools类中方法主要针对Excel表格(目前仅支持：xlsx)进行处理。
    AutoDev_OtherTools类中方法为各个基础方法，为AutoDev项目的多用途工具类，其中包括了格式转换、数据标准化处理、文件保存等方法，可以根据任务特性直接调用。
    各类中方法可以查看方法名下的详细介绍(填坑ing)。

    This Python file contains the AutoDev_SheetTools class and the AutoDev_OtherTools class.
    When instantiating objects, they are typically abbreviated as ADST and ADOT.

    Methods in the AutoDev_SheetTools class are primarily designed for processing Excel files (currently supporting only: xlsx format).

    Methods in the AutoDev_OtherTools class are fundamental utility functions serving multiple purposes across the AutoDev project. These include format conversion, data standardization, file saving, and other common operations, which can be directly called according to task requirements.

    Detailed descriptions of each method can be found in the documentation below the respective method names (documentation in progress).

    注意：
    使用os.getcwd()获取的目录情况如：/home/py-auto-dev/AutoDevPro，会获取到该项目文件下。


"""

""" ↓↓ 这是方法分类注释 ↓↓ """
""" 上下注释中间内容为某个分类的方法 """
""" ↑↑ 这是方法分类注释 ↑↑ """

"""----- AutoDev_SheetTools类头部 -----"""
class AutoDev_SheetTools:  
    def __init__(self):
        self.ADOT = AutoDev_OtherTools()

    def ADST_get_sheet_names(self,file_path):
        try:
            # 使用 pandas 的 ExcelFile 获取工作表名称
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names  # 返回所有 sheet 名称的列表
            return sheet_names
        except Exception as e:
            print(f"读取 Excel 文件时发生错误：{e}")
            return None

    def ADST_Get_Sheet_rows(self,file_path,sheet_name):
        """
        取出表格中的所有行，取出的数据type为列表嵌套元组，每一行一个元组
        """
        # 加载工作簿和工作表
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        # 获取所有行数据
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            print("工作表为空")
            return []

        else:
            return rows

    def ADST_Get_InitConfig(self):
        # 构造InitConfig文件的文件路径
        InitConfigFile_Path = os.getcwd() + "/AutoDevProFile/Temporary/Temporary_InitConfig.json"
        # 从json文件中读取InitConfig信息
        Device_List = self.ADOT.ADOT_Read_Data_From_json(InitConfigFile_Path)
        # 将FTP_Server的条目删除
        Device_List = self.ADOT.ADOT_InputList_Deletedict_value(Device_List,key="FTP_Server",value=1)
        DeviceName_List = self.ADOT.ADOT_InputList_Getdict_value(Device_List,key="Device_Name")
        return Device_List,DeviceName_List

    def ADST_Export_rows_Standardization(self,Sheet_rows):
        """
        传入行数据，将第一行作为键，并将其他行作为值，输出列表。
        """
        # 第一行作为键（列名）
        headers = [cell.strip() if cell else "" for cell in Sheet_rows[0]]
        # 构建结果列表
        Standardization_List = []
        for row in Sheet_rows[1:]:
            # 每一行构建一个字典，与列头对应
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            Standardization_List.append(row_dict)

        return Standardization_List
        
    """ ↓↓ 该分类框中方法用于将表格数据标准化并保存为配置文件 ↓↓ """
    
    def ADST_Export_Init_Sheet_dict(self,file_path):
        """
        该方法用于导出Init_Sheet中的信息并进行标准化
        """
        sheet_name = "Init_Sheet"
        Init_Sheet_rows = self.ADST_Get_Sheet_rows(file_path,sheet_name)
        Init_Sheet_Standardization_List = self.ADST_Export_rows_Standardization(Init_Sheet_rows)
        return Init_Sheet_Standardization_List

    def ADST_Export_Sheet_Standardization_dict(self,file_path,sheet_name):
        Sheet_rows = self.ADST_Get_Sheet_rows(file_path,sheet_name)
        Sheet_Standardization_List = self.ADST_Export_rows_Standardization(Sheet_rows)
        return Sheet_Standardization_List

    def ADST_Init_Sheet_dict_Save_as_json_temp(self,file_path):
        """
        该方法用于将Init_Sheet中标准化好的信息保存为一个临时文件。
        指定保存目录：./Device_Config_Backup/Temporary/
        指定命名为：Temporary_InitConfig.json
        """
        data = self.ADST_Export_Init_Sheet_dict(file_path)
        self.ADOT.ADOT_Data_Tran_File(data,file_name="InitConfig",save_dir="AutoDevProFile/Temporary/",file_format="json")
        
    def ADST_Sheet_dict_Save_as_json_temp(self,file_path,sheet_names):
        """
        该方法用于将表格中标准化好的信息保存为一个临时文件。
        指定保存目录：./Device_Config_Backup/Temporary/
        指定命名为：Temporary_InitConfig.json
        该方法默认进行批处理。
        """
        if type(sheet_names) == list:
            for sheet_name in sheet_names:
                data = self.ADST_Export_Sheet_Standardization_dict(file_path,sheet_name)
                self.ADOT.ADOT_Data_Tran_File(data,file_name=sheet_name,save_dir="AutoDevProFile/Temporary/",file_format="json")
                print(f"{sheet_name}表格已标准化处理并存储为临时文件")
                
        elif type(sheet_names) == str:
            data = self.ADST_Export_Sheet_Standardization_dict(file_path,sheet_names)
            self.ADOT.ADOT_Data_Tran_File(data,file_name=sheet_names,save_dir="AutoDevProFile/Temporary/",file_format="json")
            print(f"{sheet_names}表格已标准化处理并存储为临时文件")
            
        else:
            print(f"你传入的sheet_names为：{type(sheet_names)}类型，该方法传入参数需要为字符串或列表类型。")    
    
    def ADST_Config_Classify_By_Device(self,sheet_names):
        """
        该方法用于将从表格中标准化好导出的文件变成符合配置使用的配置文件。                 
        """
        DeviceName_List = self.ADST_Get_InitConfig()
        # print(DeviceName_List)
        for DeviceName in DeviceName_List:
            Save_List = [{"Device_Name":DeviceName}]
            # print(type(DeviceName))
            if type(sheet_names) == list:
                for sheet_name in sheet_names:
                    ConfigSheet_Path = os.getcwd() + "/AutoDevProFile/Temporary/Temporary_" + sheet_name + ".json"
                    Sheet_List = self.ADOT.ADOT_Read_Data_From_json(ConfigSheet_Path)
                    
                    Output_Config_dict = self.ADOT.ADOT_InputList_FindDictByValue(Sheet_List,key="Device_Name", value=DeviceName)
                    Output_Config_dict = self.ADOT.ADOT_InputList_DeleteKeyFromDict(Output_Config_dict,key="Device_Name")
                    # 去除设备标签
                    
                    # print(Output_Config_dict)
                    Save_List_Sub = [sheet_name] + Output_Config_dict
                    # print(Save_List_Sub)
                    for item in Output_Config_dict:
                        Save_List.append({
                            "sheet_name": sheet_name,
                            "config": item
                        })
                # print(Save_List)
                self.ADOT.ADOT_Data_Tran_File(Save_List,file_name=DeviceName,file_format="json")         
                print(f"{sheet_name}表格已标准化处理并存储为临时文件")
                
            elif type(sheet_names) == str:
                for DeviceName in DeviceName_List:
                    Save_List = [{"Device_Name":DeviceName}]

                    ConfigSheet_Path = os.getcwd() + "/AutoDevProFile/Temporary/Temporary_" + sheet_names + ".json"
                    Sheet_List = self.ADOT.ADOT_Read_Data_From_json(ConfigSheet_Path)
                    
                    Output_Config_dict = self.ADOT.ADOT_InputList_FindDictByValue(Sheet_List,key="Device_Name", value=DeviceName)
                    Output_Config_dict = self.ADOT.ADOT_InputList_DeleteKeyFromDict(Output_Config_dict,key="Device_Name")                
                    # print(Output_Config_dict)
                    for item in Output_Config_dict:
                        Save_List.append({
                            "sheet_name": sheet_names,
                            "config": item
                        })
                    # print(Save_List)
                    self.ADOT.ADOT_Data_Tran_File(Save_List,file_name=DeviceName,file_format="json")         
                print(f"{sheet_names}表格已标准化处理并存储为临时文件")
                     
            else:
                print(f"你传入的sheet_names为：{type(sheet_names)}类型，该方法传入参数需要为字符串或列表类型。") 

    def ADST_Get_Standardization_Config_list(self,Config_list):
        if not Config_list:
            return []

        sheet_name = Config_list[0]['sheet_name']
        Standardization_Config_list = [{'sheet_name': sheet_name}]

        for item in Config_list:
            Standardization_Config_list.append(item['config'])
    
        return Standardization_Config_list
    
    """ ↑↑ 该分类框中方法用于将表格数据标准化并保存为配置文件 ↑↑ """


    """ ↓↓ 该分类框中方法用于抽取配置文件中的对应部分 ↓↓ """
    def ADST_GetConfig(self,DeviceName,sheet_name):

        ConfigFile_Path = os.getcwd() + "/AutoDevProFile/Temporary/Temporary_" + DeviceName + ".json"
        Config_List = self.ADOT.ADOT_Read_Data_From_json(ConfigFile_Path)
        # print(Config_List)

        Config_List_sub = self.ADOT.ADOT_InputList_FindDictByValue(Config_List,key="sheet_name",value=sheet_name)
        
        return Config_List_sub
        

        
    
    
    """ ↑↑ 该分类框中方法用于抽取配置文件中的对应部分 ↑↑ """

    """ ↓↓ 该分类框中方法用于特殊用途 ↓↓ """

    
    """ ↑↑ 该分类框中方法用于特殊用途 ↑↑ """

"""----- AutoDev_SheetTools类结尾 -----"""

"""----- AutoDev_OtherTools类头部 -----"""
class AutoDev_OtherTools:
    def __init__(self):
        pass

    """ ↓↓ 该分类框中方法用于获取或生成各类基础信息(如：路径、文件名等) ↓↓ """
    
    def ADOT_GetDate_FileName(self,device_name,original_filename):
        """
        获取当前时间，并赋予文件名，文件名组成格式：device_name_datetime_original_filename
        """
        timestamp = datetime.now()
        date_filename = f"{device_name}_{timestamp.year}_{timestamp.month}_{timestamp.day}_{timestamp.hour}_{timestamp.minute}_{original_filename}"
        return date_filename

    """
    废弃方法
    已将功能集成到ADOT_Get_FileName_Temporary方法中，通过Date_Name是否为True，输出带时间的文件名。
    """
    # def ADOT_GetDate_FileName_Temporary(self,original_filename,file_title="Temporary"):
    #     timestamp = datetime.now()
    #     date_filename_temp = f"{file_title}_{timestamp.year}_{timestamp.month}_{timestamp.day}_{timestamp.hour}_{timestamp.minute}_{original_filename}"
    #     return date_filename_temp

    def ADOT_Get_FileName_Temporary(self, original_filename, file_title="Temporary", include_date=False):
        """
        生成临时文件名。
    
        :param original_filename: 原始文件名
        :param file_title: 文件前缀，默认为 "Temporary"
        :param include_date: 是否包含日期时间戳
        :return: 生成的文件名
        """
        if not include_date:
            return f"{file_title}_{original_filename}"
    
        now = datetime.now()
        date_str = now.strftime("%Y_%m_%d_%H_%M")  # 年_月_日_时_分，自动补零
        return f"{file_title}_{date_str}_{original_filename}"
    
    def ADOT_get_last_line(self, output):
        lines = output.strip().splitlines()
        for line in reversed(lines):
            if line.strip():
                return line.strip()
        return ""
    
    def ADOT_GetAndCreat_contents(self,folder):
        """获取当前目录，并构造文件目录路径"""
        try:
            # 获取当前 .py 文件所在的目录
            current_dir = os.getcwd()
            
            # 构造文件目录的路径
            folder_dir = os.path.join(current_dir, folder)

            return folder_dir

        except Exception as e:
        # 捕获异常并打印错误信息
            print(f"An error occurred: {e}")
            return False

    def ADOT_Get_Contents(self):
        """该方法仅用于测试"""
        current_dir = os.getcwd()
        print(current_dir)

    
    """ ↑↑ 该分类框中方法用于获取或生成各类基础信息(如：路径、文件名等) ↑↑ """

    
    """ ↓↓ 该分类框中方法用于处理各类检查任务 ↓↓ """
    
    def ADOT_Check_File(self,target_file_path):
        try:   
            # 检查文件是否存在
            if os.path.isfile(target_file_path):
                return True
            else:
                # 文件不存在时返回 False 并打印错误信息
                print(f"Error: File '{target_file_path}' does not exist.")
                return False
        except Exception as e:
        # 捕获异常并打印错误信息
            print(f"An error occurred: {e}")
            return False

    def ADOT_CheckEx_File_Or_Folder(self, target_path):
        """
        检查目标文件是否存在，并验证其所在目录是否可写（如用于生成文件）
        """
        try:
            # 检查文件是否存在
            if os.path.isfile(target_path):
                print(f"File exists: {target_path}")
                return True

            # 文件不存在，检查所在目录
            file_dir = os.path.dirname(target_path)

            # 如果路径是纯文件名（如 data.json），默认当前目录
            if not file_dir:
                file_dir = "."

            if os.path.isdir(file_dir):
                print(f"File does not exist, but directory is valid: {file_dir}")
            else:
                print(f"Error: Directory does not exist: {file_dir}")
                return False  # 目录都不存在，无法创建文件

            return False  # 文件不存在，但目录存在（可选是否返回 True？看需求）

        except Exception as e:
            print(f"An error occurred: {e}")
            return False
            
    def ADOT_Check_All_Value_Equal(self, dict_list, key, target_value):
        """
        检查字典列表中，每个字典在指定 key 上的值是否都等于 target_value。

        :param dict_list: 字典组成的列表，例如: [{"ip": "192.168.1.1", "reachable": True}, ...]
        :param key: 要检查的键名，如 "reachable"
        :param target_value: 期望的值，如 True
        :return: (bool, list) 是否全部匹配，以及不匹配的项（可用于排查）
        """
        if not dict_list:
            return False, []  # 空列表视为不满足

        unmatched = []
        all_equal = True

        for item in dict_list:
            if key not in item:
                unmatched.append(item)
                all_equal = False
            elif item[key] != target_value:
                unmatched.append(item)
                all_equal = False

        return all_equal, unmatched

    def ADOT_Check_IP_And_Mask_Sparate(self,IP_And_Mask):
        """
        该方法用于检查IP/MASK的CIDR格式是否为空，如非空则将其处理为IP +  MASK，并返回True
        参数:
            IP_And_Mask (str): 可能为空或形如 "10.10.11.2/24" 的字符串

        返回:
            result: None 表示失败；否则返回任意非 None 值（比如 True）
            Port_IP: 拆分后的 IP 地址（str），如 "10.10.11.2"
            IP_Mask: 子网掩码（str），如 "255.255.255.0"

        示例：
            ADOT_Check_IP_And_Mask_Sparate("10.10.11.2/24") -> (True, "10.10.11.2", "255.255.255.0")
            ADOT_Check_IP_And_Mask_Sparate("") -> (None, None, None)
            ADOT_Check_IP_And_Mask_Sparate("invalid") -> (None, None, None)
        """
        if not IP_And_Mask or not IP_And_Mask.strip():
            return None, None, None

        try:
            # 去除首尾空格
            ip_mask_str = IP_And_Mask.strip()

            # 使用 ipaddress 解析 CIDR 格式
            net = ipaddress.IPv4Interface(ip_mask_str)

            ip_addr = str(net.ip)        # 主机 IP（如 10.10.11.2）
            netmask = str(net.netmask)   # 子网掩码（如 255.255.255.0）

            return True, ip_addr, netmask

        except (ValueError, ipaddress.AddressValueError, ipaddress.NetmaskValueError):
            # 任何格式错误都返回 None
            return None, None, None
    

    """ ↑↑ 该分类框中方法用于处理各类检查任务 ↑↑ """

    
    """ ↓↓ 该分类框中方法用于处理各类数据格式间的转换 ↓↓ """
    
    def ADOT_list_to_string(self,input_list):
        """
        将任意类型的列表转换为用空格分隔的字符串
        """
        return ' '.join(str(item) for item in input_list)

    def ADOT_InputList_Getdict_value(self,input_list,key):
        """返回标准化里边中的特殊键的值(列表输出)"""
        return [device[key] for device in input_list if key in device]

    def ADOT_InputList_Deletedict_value(self,input_list,key,value):
        """删除特定键中包含特定值的字典"""
        return [item for item in input_list if item.get(key) != value]

    def ADOT_InputList_DeleteKeyFromDict(self, input_list, key):
        """
        从列表中每个字典中删除指定的键（key），如果该键存在
        :param input_list: 包含多个字典的列表
        :param key: 要删除的键（如 'Device_Name'）
        :return: 删除指定键后的新列表（原列表不修改）
        """
        new_list = []
        for item in input_list:
            new_item = item.copy()  # 创建副本避免修改原字典
            new_item.pop(key, None)  # 删除指定键，如果不存在也不报错
            new_list.append(new_item)
        return new_list
    
    def ADOT_InputList_FindDictByValue(self, input_list, key, value):
        """
        返回 input_list 中键为 key 且其值等于 value 的所有字典。
        返回结果是一个列表，可能包含多个字典，如果没有匹配项则返回空列表。

        :param input_list: 包含字典的列表
        :param key: 要查找的字典键
        :param value: 要匹配的值
        :return: 匹配条件的字典列表
        """
        return [item for item in input_list if item.get(key) == value]

    def ADOT_InputList_ToDeviceIPDict(self, input_list,key1,key2):
        """
        将包含设备信息的列表转换为 {Device_Name: Manage_IP} 的字典。

        :param input_list: 包含字典的列表，每个字典应包含 'Device_Name' 和 'Manage_IP' 键
        :return: 字典，格式为 {"Device_Name": "Manage_IP", ...}
        """
        return {item[key1]: item[key2] for item in input_list if key1 in item and key2 in item}

    def ADOT_Dict_ToKeyListValueList_Sorted(self, input_dict):
        """
        按键（Device_Name）排序后提取键和值列表
        """
        sorted_items = sorted(input_dict.items())  # 按键排序
        key_list = [item[0] for item in sorted_items]
        value_list = [item[1] for item in sorted_items]
        return key_list, value_list

    def ADOT_RemoveValues_From_json(file_path,value=None):
        """
        读取 JSON 文件，删除每个字典中值为 None 的键值对
        :param file_path: 输入的 JSON 文件路径
        :param output_path: 处理后的输出文件路径（可选）
        :return: 处理后的数据列表
        """
        try:
            # 1. 读取 JSON 文件
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        
            print("原始数据读取成功")

            # 2. 遍历列表中的每个字典，删除值为 None 的键值对
            cleaned_data = []
            for item in data:
                if isinstance(item, dict):
                    cleaned_item = {k: v for k, v in item.items() if v != value}
                    cleaned_data.append(cleaned_item)
                else:
                    # 如果不是字典，原样保留（一般不会）
                    cleaned_data.append(item)

            return cleaned_data

        except FileNotFoundError:
            print(f"错误：找不到文件 {file_path}")
        except json.JSONDecodeError as e:
            print(f"JSON 解析错误: {e}")
        except Exception as e:
            print(f"发生未知错误: {e}")

    
    
    """ ↑↑ 该分类框中方法用于处理各类数据格式间的转换 ↑↑ """

    
    """ ↓↓ 该分类框中方法用于处理各类文件的读写(包含txt、json、xlsx和csv) ↓↓ """
    
    def ADOT_Save_as_txt(self, data, file_path):
        """保存为 .txt 文件"""
        content = data if isinstance(data, str) else str(data)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"TXT 文件已保存: {file_path}")

    def ADOT_Save_as_json(self, file_path, data=None):
        """保存为 .json 文件"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"JSON 文件已保存: {file_path}")

    def ADOT_Save_as_csv(self, data, file_path):
        """保存为 .csv 文件"""
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            if isinstance(data, list):
                if len(data) > 0 and isinstance(data[0], dict):
                    # 字典列表：写 header 和 rows
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
                else:
                    # 二维列表或普通列表
                    for row in data:
                        if isinstance(row, (list, tuple)):
                            writer.writerow(row)
                        else:
                            writer.writerow([row])  # 单个元素转为列表
            else:
                writer.writerow([])  # 空数据
        print(f"CSV 文件已保存: {file_path}")
    
    def ADOT_Save_as_xlsx(self, data, file_path):
        """保存为 .xlsx 文件"""
        if isinstance(data, pd.DataFrame):
            df = data
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame([data])  # 其他类型转为单行 DataFrame
        df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"XLSX 文件已保存: {file_path}")

    def ADOT_Read_Data_From_json(self, file_path):
        """
        从 JSON 文件中读取设备列表

        :param file_path: JSON 文件的路径
        :return: 解析后的设备列表（list of dicts）
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        except FileNotFoundError:
            print(f"错误：文件 {file_path} 未找到。")
            return []
        except json.JSONDecodeError as e:
            print(f"错误：JSON 解析失败 - {e}")
            return []
        except Exception as e:
            print(f"发生未知错误：{e}")
            return []

    """ ↑↑ 该分类框中方法用于处理各类文件的读写(包含txt、json、xlsx和csv) ↑↑ """


    """ ↓↓ 该分类框中方法作用请查看方法名下详细说明 ↓↓ """
    def ADOT_Data_Tran_File(self,Data,file_name="Unknown_File",save_dir="AutoDevProFile/Temporary",file_format="txt",include_date=False):
        """
        将数据保存为指定格式的文件（txt 或 json），并存储到指定目录。

        :param data: 要保存的数据（str, dict, list 等）
        :param file_name: 文件名（不带扩展名）
        :param save_dir: 保存目录路径
        :param file_format: 文件格式，'txt' 或 'json'
        """
        
        file_name = self.ADOT_Get_FileName_Temporary(original_filename=file_name,include_date=include_date)

        # 检查格式是否合法
        if file_format not in ['txt', 'json','csv','xlsx']:
            raise ValueError("file_format must be 'txt' ,'csv' ,'xlsx' or 'json'")

        # 确保保存目录存在
        os.makedirs(save_dir, exist_ok=True)

        # 构造完整文件路径
        file_path = os.path.join(save_dir, f"{file_name}.{file_format}")

        # 分发处理
        try:
            if file_format == 'txt':
                self.ADOT_Save_as_txt(Data, file_path)
            elif file_format == 'json':
                self.ADOT_Save_as_json(file_path, Data)
            elif file_format == 'csv':
                self.ADOT_Save_as_csv(Data, file_path)
            elif file_format == 'xlsx':
                self.ADOT_Save_as_xlsx(Data, file_path)
 
            print(f"✅ 已成功存储文件到：{file_path}")

            return True  # 返回文件路径表示成功

        except Exception as e:
            print(f"保存文件失败: {e}")
            raise  # 可选：重新抛出异常供上层捕获

    """ ↑↑ 该分类框中方法作用请查看方法名下详细说明 ↑↑ """

    
    """ ↓↓ 该分类框中方法用于特殊用途 ↓↓ """
    
    def ADOT_Createdict_passresult(self):
        """
        该方法用于生成测试通过的结果，并附当前时间。
        """
        now = datetime.now()
        test_result = {"result": "pass", "datetime": now.isoformat()}
        return test_result
        
    """ ↑↑ 该分类框中方法用于特殊用途 ↑↑ """
    
    
"""----- AutoDev_OtherTools类结尾 -----"""       


